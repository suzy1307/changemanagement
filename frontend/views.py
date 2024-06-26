import base64
import os
import json
from io import BytesIO
import time
import openpyxl

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.views import View
from django import forms
from django.conf import settings
import requests
import openai

from .forms import InputForm, DynamicForm
from datasets import load_dataset

PROJECTS = ['PLDT', 'Maxis', 'TKS']

class AboutView(View):
    template_name = 'input_form.html'

    def get(self, request, *args, **kwargs):
        form = InputForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = InputForm(request.POST)
        if form.is_valid():
            data = {
                "url": form.cleaned_data['url'],
                "credentials": {
                    "username": form.cleaned_data['username'],
                    "password": form.cleaned_data['password']
                },
                "search_term": form.cleaned_data['search_term'],
                "summary_text": form.cleaned_data['summary_text'],
                "affected_area": form.cleaned_data['affected_area'],
                "company_name": form.cleaned_data['company_name'],
                "requested_date": form.cleaned_data['requested_date'].strftime('%b %d, %Y'),
                "end_date": form.cleaned_data['end_date'].strftime('%b %d, %Y'),
                "Number_of_Users_selector_": form.cleaned_data['number_of_users'],
                "Affected_Services_selector_": form.cleaned_data['affected_services'],
                "Resources_required_to_Prepare_selector_": form.cleaned_data['resources_required'],
                "Preparation_Effort_selector_": form.cleaned_data['preparation_effort'],
                "Implementation_duration_selector_": form.cleaned_data['implementation_duration'],
                "Failure_Exposure_selector_": form.cleaned_data['failure_exposure'],
            }
            response = HttpResponse(json.dumps(data), content_type='application/json')
            response['Content-Disposition'] = 'attachment; filename="data.json"'
            return response
        return render(request, self.template_name, {'form': form})

class ProjectSelectForm(forms.Form):
    project = forms.ChoiceField(choices=[(project, project) for project in PROJECTS], label="Select Project")

class ProjectSelectView(View):
    template_name = 'project_select.html'
    form_class = ProjectSelectForm

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            project = form.cleaned_data['project']
            request.session['project_name'] = project
            return redirect('fill_project_template_step1', project_name=project)
        return render(request, self.template_name, {'form': form})



from django.shortcuts import redirect

class FillProjectTemplateStep1View(View):
    template_name = 'fill_project_template_step1.html'

    def get_placeholder_fields(self, wb):
        placeholders = set()
        for sheet in wb:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        if '<' in cell.value and '>' in cell.value:
                            placeholders.add(cell.value.strip('<>'))
        return placeholders

    def get(self, request, project_name, *args, **kwargs):
        template_path = os.path.join(settings.BASE_DIR, 'templates', f'{project_name}.xlsx')
        try:
            wb = openpyxl.load_workbook(template_path)
        except FileNotFoundError:
            return HttpResponse(f'Template file not found: {template_path}', status=404)
        except openpyxl.utils.exceptions.InvalidFileException:
            return HttpResponse(f'Invalid template file: {template_path}', status=400)

        placeholders = self.get_placeholder_fields(wb)

        # Dynamically create a form with fields corresponding to placeholders
        form_class = type('DynamicForm', (DynamicForm,), {placeholder: forms.CharField(label=placeholder) for placeholder in placeholders})
        form = form_class()

        return render(request, self.template_name, {'form': form, 'project_name': project_name})

    def post(self, request, project_name, *args, **kwargs):
        template_path = os.path.join(settings.BASE_DIR, 'templates', f'{project_name}.xlsx')
        try:
            wb = openpyxl.load_workbook(template_path)
        except FileNotFoundError:
            return HttpResponse(f'Template file not found: {template_path}', status=404)
        except openpyxl.utils.exceptions.InvalidFileException:
            return HttpResponse(f'Invalid template file: {template_path}', status=400)

        placeholders = self.get_placeholder_fields(wb)

        # Dynamically create a form with fields corresponding to placeholders
        form_class = type('DynamicForm', (DynamicForm,), {placeholder: forms.CharField(label=placeholder) for placeholder in placeholders})
        form = form_class(request.POST)

        if form.is_valid():
            try:
                for sheet in wb:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str):
                                if '<' in cell.value and '>' in cell.value:
                                    placeholder = cell.value.strip('<>')
                                    if placeholder in form.cleaned_data:
                                        cell.value = form.cleaned_data[placeholder]

                # Save the filled workbook to BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                # # Prepare HTTP response with the filled Excel file
                # response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                # response['Content-Disposition'] = f'attachment; filename="{project_name}_filled.xlsx"'
                # return response

            except Exception as e:
                return HttpResponse(f'Error while processing the workbook: {str(e)}', status=500)

        return redirect('fill_project_template_step2', project_name=project_name)


class FillProjectTemplateStep2View(View):
    def get_placeholder_fields(self, wb):
        placeholders = set()
        for sheet in wb:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and '**' in cell.value:
                        placeholders.add(cell.value.strip('**'))
        return placeholders

    def send_to_private_llm(self, text):
        url = "https://chat-ms-amaiz-backend-ms-prod.at-azure-amaiz-55185456.corp.amdocs.azr/api/v1/chats/send-message"
        headers = {
            "accept": "application/json",
            "Content-Type": "application/json"
        }
        data = {
            "username": "shwetac",  # Your username
            "apikey": "01923c96-48f8-4fd4-9aae-651e3bee9586",  # Your API key
            "conv_id": "",
            "messages": [
                {
                    "user": text  # Use the text parameter here
                }
            ],
            "promptfilename": "",
            "promptname": "",
            "prompttype": "system",
            "promptrole": "act as ChatGPT",
            "prompttask": "",
            "promptexamples": "",
            "promptformat": "",
            "promptrestrictions": "",
            "promptadditional": "",
            "max_tokens": 4000,
            "model_type": "GPT3.5_16K",
            "temperature": 0.1,
            "topKChunks": 2,
            "read_from_your_data": False,
            "document_groupname": "",
            "document_grouptags": [],
            "data_filenames": [],
            "find_the_best_response": False,
            "chat_attr": {},
            "additional_attr": {}
        }

        ca_bundle_path = r'C:\Users\shwetac\project\changemanagement\certs\amdcerts.pem'  # Path to your CA bundle

        try:
            response = requests.post(url, headers=headers, data=json.dumps(data), verify=ca_bundle_path)
            response.raise_for_status()  # Raise an exception for HTTP errors
            response_json = response.json()
            task_id = response_json.get("task_id")
            if not task_id:
                raise RuntimeError("Error: No task_id in response")

            status_url = f"https://chat-ms-amaiz-backend-ms-prod.at-azure-amaiz-55185456.corp.amdocs.azr/api/v1/chats/status/{task_id}"
            status_headers = {'accept': 'application/json'}

            # Poll the status endpoint until the task is complete
            result = ""
            while True:
                chat_results = requests.get(status_url, headers=status_headers, verify=ca_bundle_path)
                chat_data = chat_results.json()
                if chat_data.get("status") == "Complete":
                    result = chat_data.get("result")
                    break
                time.sleep(1)

            return result

        except requests.RequestException as e:
            print(f"Error communicating with private LLM API: {str(e)}")  # Debugging: Log the error
            raise RuntimeError(f'Error communicating with private LLM API: {str(e)}')

    def load_workbook(self, template_path):
        try:
            return openpyxl.load_workbook(template_path)
        except FileNotFoundError:
            raise FileNotFoundError(f'Template file not found: {template_path}')
        except openpyxl.utils.exceptions.InvalidFileException:
            raise ValueError(f'Invalid template file: {template_path}')

    def create_dynamic_form(self, placeholders):
        return type('DynamicForm', (forms.Form,), {placeholder: forms.CharField(label=placeholder) for placeholder in placeholders})

    def get(self, request, project_name):
        template_path = os.path.join(settings.BASE_DIR, 'templates', f'{project_name}.xlsx')
        try:
            wb = self.load_workbook(template_path)
        except (FileNotFoundError, ValueError) as e:
            return HttpResponse(str(e), status=404 if isinstance(e, FileNotFoundError) else 400)

        placeholders = self.get_placeholder_fields(wb)
        form_class = self.create_dynamic_form(placeholders)
        form = form_class()

        return render(request, 'fill_project_template_step2.html', {'form': form, 'project_name': project_name})

    def post(self, request, project_name):
        template_path = os.path.join(settings.BASE_DIR, 'templates', f'{project_name}.xlsx')
        try:
            wb = openpyxl.load_workbook(template_path)
        except FileNotFoundError:
            return JsonResponse({'error': f'Template file not found: {template_path}'}, status=404)
        except openpyxl.utils.exceptions.InvalidFileException:
            return JsonResponse({'error': f'Invalid template file: {template_path}'}, status=400)

        placeholders = self.get_placeholder_fields(wb)
        form_class = self.create_dynamic_form(placeholders)
        form = form_class(request.POST)

        if form.is_valid():
            try:
                filled_wb = openpyxl.Workbook()

                for sheet in wb:
                    filled_sheet = filled_wb.create_sheet(title=sheet.title)
                    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                        for col_idx, cell in enumerate(row, start=1):
                            if isinstance(cell.value, str) and '**' in cell.value:
                                placeholder = cell.value.strip('**')
                                if placeholder in form.cleaned_data:
                                    ai_response = self.send_to_private_llm(form.cleaned_data[placeholder])
                                    filled_sheet.cell(row=row_idx, column=col_idx, value=ai_response)
                            else:
                                filled_sheet.cell(row=row_idx, column=col_idx, value=cell.value)

                filled_buffer = BytesIO()
                filled_wb.save(filled_buffer)
                filled_buffer.seek(0)

                # Prepare HttpResponse to trigger file download
                response = HttpResponse(filled_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{project_name}_filled.xlsx"'

                return response

            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)

        return JsonResponse({'errors': form.errors}, status=400)